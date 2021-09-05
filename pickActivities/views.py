from django.shortcuts import render
import openpyxl
import random
from PIL import Image
import matplotlib.pyplot as plt

def matchup(request):
    book = openpyxl.load_workbook('pickActivities/Head To Head - Austin.xlsx')
    sheet = book['Sheet1']
    
    try:
        lastRoundWinnerRowNum = int(request.GET['winner'])
        sheet.cell(row=lastRoundWinnerRowNum,column=2).value += 1
        book.save('pickActivities/Head To Head - Austin.xlsx')
    except:
        pass
        
    optionNums = range(2,sheet.max_row+1)
    [option1RowNum,option2RowNum] = random.sample(optionNums,2)
    option1Title = sheet.cell(row=option1RowNum,column=1).value
    option2Title = sheet.cell(row=option2RowNum,column=1).value
    pic1 = Image.open('pickActivities/static/pickActivities' + option1Title + '.PNG')
    pic2 = Image.open('pickActivities/static/pickaActivities' + option2Title + '.PNG')
   
    pic1.save('pickActivities/static/pickActivities/pic1.PNG')
    pic2.save('pickActivities/static/pickActivities/pic2.PNG')
    context = {'option1Title':option1Title, 'option2Title':option2Title, 'option1RowNum':option1RowNum, 'option2RowNum':option2RowNum}
    return render(request, 'pickActivities/matchup.html', context)



def results(request):
    book = openpyxl.load_workbook('pickActivities/Head To Head - Austin.xlsx')
    sheet = book['Sheet1']  
    plt.clf()  

    optionTitles = []
    battlesWon = []
    for rowNum in range(2,sheet.max_row+1):
        optionTitles += [sheet.cell(row=rowNum,column=1).value]
        battlesWon += [sheet.cell(row=rowNum,column=2).value]

    fig, ax = plt.subplots(figsize = (8, 5))
    ax.barh(optionTitles,battlesWon)
    ax.invert_yaxis()
    ax.set_title('What Is The Best Thing To Do In Austin?')
    plt.savefig('pickActivities/static/pickActivities/graph.PNG', bbox_inches="tight")

    return render(request, 'pickActivities/results.html')